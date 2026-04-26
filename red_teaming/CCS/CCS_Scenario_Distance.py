#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Scenario-cluster cosine similarity across N = 2..5 (S5, within-case).

For each case (e.g. T1-C1):
  - Each scenario = cluster of 5 traces → centroid (mean of trace embeddings)
  - For N in {2, 3, 4, 5}: take first N scenarios' centroids (S1..SN) and
    compute pairwise cosine similarity → C(N,2) pairs per case
  - Average within the case (case_mean)
  - Macro-average across cases (overall_mean)

Usage:
    python CCS_Scenario_Cluster_Distance.py                      # all agents
    python CCS_Scenario_Cluster_Distance.py --agent banking_cs_agent
"""

from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from itertools import combinations
from pathlib import Path

import numpy as np


DEFAULT_SCENARIOS_ROOT = (
    r"C:\Users\최유진\Desktop\VSCode\Agent_AI_Security"
    r"\red_teaming\CCS\generated_scenarios"
)
SCENARIO_ID_RE = re.compile(r"^(T\d+)-(C\d+)-S(\d+)$")
N_VALUES = (2, 3, 4, 5)


def discover_agents(cache_dir: Path) -> list:
    agents = []
    for f in sorted(cache_dir.glob("trace_emb_*.npz")):
        name = f.stem
        if name.startswith("trace_emb_"):
            agents.append(name[len("trace_emb_"):])
    return agents


def cosine_sim(a, b):
    denom = (np.linalg.norm(a) * np.linalg.norm(b)) + 1e-12
    return float(np.dot(a, b) / denom)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--agent", default=None)
    ap.add_argument("--cache-dir", default=".")
    ap.add_argument("--scenarios-root", default=DEFAULT_SCENARIOS_ROOT)
    ap.add_argument("--output-dir", default=None)
    ap.add_argument("--combined-xlsx", default=None)
    args = ap.parse_args()

    cache_dir = Path(args.cache_dir)

    if args.agent:
        agents = [args.agent]
    else:
        agents = discover_agents(cache_dir)
        if not agents:
            raise RuntimeError(f"No trace_emb_*.npz files in {cache_dir}")
        print(f"[auto] Processing all agents: {agents}\n")

    all_results = []
    for agent in agents:
        try:
            result = process_agent(agent, args)
            if result is not None:
                all_results.append(result)
        except Exception as e:
            print(f"[error] {agent}: {e}\n")

    if len(agents) > 1 and all_results:
        try:
            save_combined_xlsx(all_results, args)
        except Exception as e:
            print(f"[warn] combined xlsx failed: {e}")


def process_agent(agent: str, args):
    cache_path = Path(args.cache_dir) / f"trace_emb_{agent}.npz"
    if not cache_path.exists():
        raise RuntimeError(f"cache not found: {cache_path}")

    print(f"{'#' * 70}")
    print(f"# Agent: {agent}")
    print(f"{'#' * 70}")

    cached = np.load(cache_path, allow_pickle=True)
    all_keys = cached["trace_keys"].tolist()
    all_emb = cached["embeddings"]

    # Filter to S5, group by scenario_id
    s5_prefix = f"{agent}|S5|"
    by_scenario = defaultdict(list)
    for i, key in enumerate(all_keys):
        if not key.startswith(s5_prefix):
            continue
        parts = key.split("|")
        if len(parts) != 4:
            continue
        _, _, sid, _ = parts
        by_scenario[sid].append(all_emb[i])

    if not by_scenario:
        raise RuntimeError("No S5 traces in cache")

    centroids = {sid: np.mean(np.stack(vs), axis=0)
                  for sid, vs in by_scenario.items()}

    # Group scenarios by case
    by_case = defaultdict(list)
    for sid in centroids:
        m = SCENARIO_ID_RE.match(sid)
        if not m:
            continue
        t, c, s_num = m.group(1), m.group(2), int(m.group(3))
        by_case[f"{t}-{c}"].append((s_num, sid))
    for ck in by_case:
        by_case[ck].sort()

    print(f"  S5 scenarios: {len(centroids)}  |  cases: {len(by_case)}\n")

    # Per-N
    per_N = {}
    for N in N_VALUES:
        per_case = {}
        pool = []
        for case_key, seq in by_case.items():
            if len(seq) < N:
                continue
            sub = seq[:N]
            sims = []
            pair_records = []
            for (sa, sid_a), (sb, sid_b) in combinations(sub, 2):
                s = cosine_sim(centroids[sid_a], centroids[sid_b])
                sims.append(s)
                pair_records.append({
                    "a": sid_a, "b": sid_b, "cosine_similarity": s,
                })
            sims_arr = np.array(sims)
            per_case[case_key] = {
                "n_pairs": len(sims),
                "case_mean": float(sims_arr.mean()),
                "case_std": float(sims_arr.std(ddof=1))
                             if len(sims) > 1 else 0.0,
                "min": float(sims_arr.min()),
                "max": float(sims_arr.max()),
                "pairs": pair_records,
            }
            pool.extend(sims)

        if not per_case:
            continue

        case_means_arr = np.array([r["case_mean"] for r in per_case.values()])
        pool_arr = np.array(pool)
        pcts_case = np.percentile(case_means_arr, [5, 25, 50, 75, 95])

        per_N[N] = {
            "n_cases": len(per_case),
            "n_pairs_per_case": N * (N - 1) // 2,
            "n_pairs_total": len(pool_arr),
            "overall_mean": float(case_means_arr.mean()),
            "across_case_std": float(case_means_arr.std(ddof=1))
                                 if len(case_means_arr) > 1 else 0.0,
            "min_case_mean": float(case_means_arr.min()),
            "max_case_mean": float(case_means_arr.max()),
            "case_means_percentiles": {
                "p5":  float(pcts_case[0]),
                "p25": float(pcts_case[1]),
                "p50": float(pcts_case[2]),
                "p75": float(pcts_case[3]),
                "p95": float(pcts_case[4]),
            },
            "per_case": per_case,
        }

    # Console
    print(f"  {'N':>2}  {'n_cases':>7}  {'pairs/case':>10}  "
          f"{'n_pairs':>7}  {'overall_mean':>12}  {'across_std':>10}  "
          f"{'min':>7}  {'max':>7}  {'p25':>7}  {'p50':>7}  {'p75':>7}")
    print("  " + "-" * 110)
    for N in sorted(per_N.keys()):
        r = per_N[N]
        pcts = r["case_means_percentiles"]
        print(f"  {N:>2}  {r['n_cases']:>7}  {r['n_pairs_per_case']:>10}  "
              f"{r['n_pairs_total']:>7}  {r['overall_mean']:>12.4f}  "
              f"{r['across_case_std']:>10.4f}  "
              f"{r['min_case_mean']:>7.4f}  {r['max_case_mean']:>7.4f}  "
              f"{pcts['p25']:>7.4f}  {pcts['p50']:>7.4f}  {pcts['p75']:>7.4f}")

    # Save
    out_dir = Path(args.output_dir) if args.output_dir \
        else Path(args.scenarios_root) / agent
    out_dir.mkdir(parents=True, exist_ok=True)

    # JSON
    json_path = out_dir / "scenario_cluster_similarity.json"
    json_path.write_text(json.dumps({
        "agent": agent,
        "per_N": {str(N): r for N, r in per_N.items()},
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nSaved JSON → {json_path}")

    # CSV (long format: N, case, a, b, sim)
    csv_path = out_dir / "scenario_cluster_similarity.csv"
    lines = ["N,case,scenario_a,scenario_b,cosine_similarity"]
    for N in sorted(per_N.keys()):
        for ck in sorted(per_N[N]["per_case"].keys(),
                          key=lambda k: (
                              int(re.match(r"T(\d+)", k).group(1)),
                              int(re.search(r"C(\d+)", k).group(1)),
                          )):
            for p in per_N[N]["per_case"][ck]["pairs"]:
                lines.append(f"{N},{ck},{p['a']},{p['b']},"
                             f"{p['cosine_similarity']:.6f}")
    csv_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Saved CSV  → {csv_path}")

    # Agent XLSX
    xlsx_path = out_dir / "scenario_cluster_similarity.xlsx"
    save_agent_xlsx(agent, per_N, xlsx_path)
    print(f"Saved XLSX → {xlsx_path}")

    # Plot
    try:
        save_plots(agent, per_N, out_dir)
    except Exception as e:
        print(f"  ⚠ plot failed: {e}")

    return (agent, per_N)


# ───────────────────── xlsx helpers ─────────────────────────────────────────

def _xlsx_styles():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    FONT_NAME = "Arial"
    return {
        "header_font":  Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11),
        "header_fill":  PatternFill("solid", start_color="2E5D8C"),
        "summary_font": Font(name=FONT_NAME, bold=True, size=10),
        "summary_fill": PatternFill("solid", start_color="D9E1F2"),
        "body_font":    Font(name=FONT_NAME, size=10),
        "case_font":    Font(name=FONT_NAME, bold=True, size=10),
        "border":       Border(
            left=Side(border_style="thin", color="BFBFBF"),
            right=Side(border_style="thin", color="BFBFBF"),
            top=Side(border_style="thin", color="BFBFBF"),
            bottom=Side(border_style="thin", color="BFBFBF"),
        ),
        "center": Alignment(horizontal="center", vertical="center"),
        "left":   Alignment(horizontal="left",   vertical="center"),
    }


def _write_n_block(ws, start_row, agent, N, r, styles):
    """Write one N block = summary row + per-case rows.
    Returns next free row."""
    summary_cells = [
        f"{agent} N={N} (overall)",
        r["n_pairs_total"],
        r["overall_mean"],
        r["across_case_std"],
        r["min_case_mean"],
        r["max_case_mean"],
    ]
    for col, v in enumerate(summary_cells, start=1):
        c = ws.cell(row=start_row, column=col, value=v)
        c.font = styles["summary_font"]
        c.fill = styles["summary_fill"]
        c.border = styles["border"]
        c.alignment = styles["left"] if col == 1 else styles["center"]
        if isinstance(v, float):
            c.number_format = "0.0000"
        elif isinstance(v, int):
            c.number_format = "0"
    start_row += 1

    for ck in sorted(r["per_case"].keys(),
                      key=lambda k: (
                          int(re.match(r"T(\d+)", k).group(1)),
                          int(re.search(r"C(\d+)", k).group(1)),
                      )):
        cr = r["per_case"][ck]
        row_values = [ck, cr["n_pairs"], cr["case_mean"],
                      cr["case_std"], cr["min"], cr["max"]]
        for col, v in enumerate(row_values, start=1):
            c = ws.cell(row=start_row, column=col, value=v)
            c.border = styles["border"]
            if col == 1:
                c.font = styles["case_font"]
                c.alignment = styles["left"]
            else:
                c.font = styles["body_font"]
                c.alignment = styles["center"]
                if isinstance(v, float):
                    c.number_format = "0.0000"
                else:
                    c.number_format = "0"
        start_row += 1
    return start_row


def save_agent_xlsx(agent, per_N, path):
    from openpyxl import Workbook
    styles = _xlsx_styles()

    wb = Workbook()
    ws = wb.active
    ws.title = "Scenario cluster sim"

    for col, h in enumerate(["Case", "n_pairs", "mean", "std", "min", "max"],
                              start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = styles["header_font"]
        c.fill = styles["header_fill"]
        c.alignment = styles["center"]
        c.border = styles["border"]

    row = 2
    for N in sorted(per_N.keys()):
        row = _write_n_block(ws, row, agent, N, per_N[N], styles)

    ws.column_dimensions["A"].width = 32
    for col in "BCDEF":
        ws.column_dimensions[col].width = 12
    ws.freeze_panes = "A2"
    wb.save(path)


def save_combined_xlsx(all_results, args):
    from openpyxl import Workbook
    styles = _xlsx_styles()

    wb = Workbook()
    ws = wb.active
    ws.title = "All agents"

    for col, h in enumerate(["Case", "n_pairs", "mean", "std", "min", "max"],
                              start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = styles["header_font"]
        c.fill = styles["header_fill"]
        c.alignment = styles["center"]
        c.border = styles["border"]

    row = 2
    for agent, per_N in all_results:
        for N in sorted(per_N.keys()):
            row = _write_n_block(ws, row, agent, N, per_N[N], styles)

    ws.column_dimensions["A"].width = 32
    for col in "BCDEF":
        ws.column_dimensions[col].width = 12
    ws.freeze_panes = "A2"

    out_path = Path(args.combined_xlsx) if args.combined_xlsx \
        else Path(args.scenarios_root) / "scenario_cluster_similarity_all.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"\n{'=' * 70}")
    print(f"Combined XLSX saved → {out_path}")


def save_plots(agent, per_N, out_dir):
    import matplotlib.pyplot as plt

    BG = "#141414"; TEXT = "#ececec"; GRID = "#2e2e2e"

    fig, ax = plt.subplots(figsize=(9, 6), facecolor=BG)
    ax.set_facecolor(BG)
    Ns = sorted(per_N.keys())
    means = [per_N[N]["overall_mean"] for N in Ns]
    stds = [per_N[N]["across_case_std"] for N in Ns]
    ax.errorbar(Ns, means, yerr=stds, fmt="o-",
                 color="#4da3ff", ecolor="#ff9966",
                 elinewidth=1.5, capsize=5, markersize=9,
                 label="overall_mean ± across_case_std")
    for N, m in zip(Ns, means):
        ax.annotate(f"{m:.4f}", xy=(N, m), xytext=(8, 8),
                     textcoords="offset points", color=TEXT, fontsize=9)
    ax.set_title(
        f"Scenario cluster cosine similarity vs N\n{agent} / S5",
        fontsize=12, color=TEXT, pad=12,
    )
    ax.set_xlabel("N (# scenarios per case)", color=TEXT)
    ax.set_ylabel("cosine similarity (macro-avg)", color=TEXT)
    ax.set_xticks(Ns)
    ax.tick_params(colors=TEXT)
    ax.grid(True, alpha=0.25, linestyle="--", color=GRID)
    ax.legend(facecolor="#1f1f1f", edgecolor="#4a4a4a",
               labelcolor=TEXT, framealpha=0.92)
    for s in ax.spines.values():
        s.set_color("#4a4a4a")
    plt.tight_layout()
    line_path = out_dir / "scenario_cluster_similarity_vs_N.png"
    plt.savefig(line_path, dpi=200, bbox_inches="tight", facecolor=BG)
    plt.close()
    print(f"Saved PNG  → {line_path}")


if __name__ == "__main__":
    main()